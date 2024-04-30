from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def dataframe_to_columns(df):
    """
    Convierte un DataFrame en una lista de listas donde cada lista interna contiene el nombre de la columna y sus valores correspondientes.

    Args:
    - df: DataFrame de Pandas

    Returns:
    - variable_value_lists: Lista de listas que contienen el nombre de la columna y sus valores
    """
    # Inicializar la lista de listas
    variable_value_lists = []

    # Iterar sobre las columnas del DataFrame
    for column_name in df.columns:
        # Obtener los valores de la columna y convertirlos en una lista
        values = df[column_name].tolist()
        # Crear una lista que contenga el nombre de la columna seguido de sus valores
        column_values_list = [column_name] + values
        # Agregar la lista de valores de la columna a la lista principal
        variable_value_lists.append(column_values_list)

    return variable_value_lists


def report_df_to_excel(df, filename, sheet_name='Sheet1', num_col=0, resize_columns=True):
    """
    Guarda un DataFrame en un archivo Excel, agregando sus columnas al final de la hoja de cálculo especificada.

    Args:
    - df: DataFrame de Pandas que se va a guardar.
    - filename: Ruta del archivo Excel donde se guardará el DataFrame.
    - sheet_name: Nombre de la hoja de cálculo en la que se agregarán los datos del DataFrame.
    - resize_columns: Booleano que indica si se deben redimensionar automáticamente las columnas para ajustarse al contenido.

    Returns:
    - None
    """

    try:
        # Intenta abrir el archivo existente
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Convierte el DataFrame en columnas y las agrega al libro de trabajo
    column_df = dataframe_to_columns(df)
    ws.append_column(column_df, column=num_col)

    # Redimensiona las columnas si se solicita
    if resize_columns:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtiene la letra de la columna
            for cell in col:
                try:  # Necesario porque algunas celdas pueden estar vacías
                    cell_value_length = len(str(cell.value))
                    if cell_value_length > max_length:
                        max_length = cell_value_length
                except:
                    pass
            # Ajusta el ancho de la columna
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    # Guarda y cierra el libro de trabajo
    wb.save(filename)
    wb.close()
